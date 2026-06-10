import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import io

st.set_page_config(page_title="Cuadrantes Metrovalencia", layout="wide")

st.title("📅 Cuadrante de Servicios - Metrovalencia")
st.caption("Vista calendario | Colores por turno | Guarda cambios en Excel")

# ============================================================
# GESTOR BD
# ============================================================

class GestorDB:
    def __init__(self, db_path="cuadrantes.db"):
        self.db_path = db_path
        self._init_db()
    
    def _init_db(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS agentes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT NOT NULL,
            nombre TEXT NOT NULL,
            zona TEXT
        )''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS turnos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            agente_id INTEGER NOT NULL,
            dia INTEGER NOT NULL,
            turno TEXT,
            UNIQUE(agente_id, dia)
        )''')
        conn.commit()
        conn.close()
    
    def limpiar(self):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("DELETE FROM turnos")
        cursor.execute("DELETE FROM agentes")
        conn.commit()
        conn.close()
    
    def guardar_agente(self, codigo, nombre, zona):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO agentes (codigo, nombre, zona) VALUES (?, ?, ?)",
            (codigo, nombre, zona)
        )
        agente_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return agente_id
    
    def guardar_turnos(self, agente_id, turnos):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        for dia, turno in enumerate(turnos, 1):
            # Limpiar turno: si es 0 o vacío lo tratamos como sin turno
            if turno and str(turno).strip() not in ["0", "0.0", ""]:
                turno_limpio = str(turno).strip()
                cursor.execute(
                    "INSERT OR REPLACE INTO turnos (agente_id, dia, turno) VALUES (?, ?, ?)",
                    (agente_id, dia, turno_limpio)
                )
            else:
                # Eliminar turno si existe y es 0
                cursor.execute("DELETE FROM turnos WHERE agente_id = ? AND dia = ?", (agente_id, dia))
        conn.commit()
        conn.close()
    
    def get_agentes(self):
        conn = sqlite3.connect(self.db_path)
        df = pd.read_sql("SELECT id, codigo, nombre, zona FROM agentes ORDER BY zona, nombre", conn)
        conn.close()
        return df
    
    def get_turnos(self, agente_id):
        conn = sqlite3.connect(self.db_path)
        df = pd.read_sql("SELECT dia, turno FROM turnos WHERE agente_id = ? ORDER BY dia", conn, params=(agente_id,))
        conn.close()
        turnos = [""] * 31
        for _, row in df.iterrows():
            turnos[row["dia"] - 1] = row["turno"]
        return turnos
    
    def get_turno(self, agente_id, dia):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT turno FROM turnos WHERE agente_id = ? AND dia = ?", (agente_id, dia))
        result = cursor.fetchone()
        conn.close()
        return result[0] if result else ""
    
    def intercambiar_turnos(self, agente1_id, agente2_id, dia):
        t1 = self.get_turno(agente1_id, dia)
        t2 = self.get_turno(agente2_id, dia)
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        # Asignar t2 a agente1 (puede ser vacío)
        if t2:
            cursor.execute(
                "INSERT OR REPLACE INTO turnos (agente_id, dia, turno) VALUES (?, ?, ?)",
                (agente1_id, dia, t2)
            )
        else:
            cursor.execute("DELETE FROM turnos WHERE agente_id = ? AND dia = ?", (agente1_id, dia))
        
        # Asignar t1 a agente2 (puede ser vacío)
        if t1:
            cursor.execute(
                "INSERT OR REPLACE INTO turnos (agente_id, dia, turno) VALUES (?, ?, ?)",
                (agente2_id, dia, t1)
            )
        else:
            cursor.execute("DELETE FROM turnos WHERE agente_id = ? AND dia = ?", (agente2_id, dia))
        
        conn.commit()
        conn.close()
        return t1, t2


# ============================================================
# RANGOS DE FILAS (FIJOS SEGÚN EXCEL ORIGINAL)
# ============================================================

RANGOS_POR_ZONA = {
    "JC": (12, 127),
    "AEZ6": (140, 181),
    "AEZ7": (196, 245),
    "AEZ8": (262, 309)
}

def cargar_por_rangos(archivo_path):
    """Carga agentes según los rangos de filas predefinidos. Filtra códigos 0 y vacantes."""
    df = pd.read_excel(archivo_path, sheet_name="MAYO 2026", header=None)
    agentes_por_zona = {}
    
    for zona, (fila_inicio, fila_fin) in RANGOS_POR_ZONA.items():
        agentes_zona = []
        for fila in range(fila_inicio - 1, min(fila_fin, len(df))):
            codigo = df.iloc[fila, 3] if df.shape[1] > 3 else None
            nombre = df.iloc[fila, 4] if df.shape[1] > 4 else None
            
            # Limpiar código
            if pd.isna(codigo):
                continue
            # Convertir a string y eliminar espacios
            codigo_str = str(codigo).strip()
            # Si es numérico 0 o string "0" -> excluir
            if codigo_str == "0" or codigo_str == "":
                continue
            
            # Limpiar nombre
            if pd.isna(nombre):
                continue
            nombre_str = str(nombre).strip()
            if "DESPLAZADO" in nombre_str.upper() or "VACANTE" in nombre_str.upper():
                continue
            
            # Leer turnos (columnas 5,7,9... hasta 5+2*30=65)
            turnos = []
            for dia in range(31):
                col_turno = 5 + (dia * 2)
                if df.shape[1] > col_turno:
                    valor = df.iloc[fila, col_turno]
                    if pd.isna(valor):
                        turnos.append("")
                    else:
                        valor_str = str(valor).strip()
                        if valor_str in ["0", "0.0"]:
                            turnos.append("")
                        else:
                            turnos.append(valor_str)
                else:
                    turnos.append("")
            
            agentes_zona.append({
                "codigo": codigo_str,
                "nombre": nombre_str,
                "turnos": turnos
            })
        agentes_por_zona[zona] = agentes_zona
    return agentes_por_zona


def guardar_cambios_en_excel(archivo_original, db):
    """
    Toma el Excel original, actualiza las celdas de turno según los datos de la BD,
    y devuelve un objeto BytesIO con el nuevo Excel.
    """
    # Cargar el libro original
    wb = openpyxl.load_workbook(archivo_original)
    ws = wb["MAYO 2026"]
    
    # Obtener todos los agentes con sus IDs y ubicaciones
    df_agentes = db.get_agentes()
    if df_agentes.empty:
        raise ValueError("No hay agentes en la BD")
    
    # Para cada agente, localizar su fila en el Excel según zona y código
    # Como el Excel tiene filas fijas por zona, necesitamos saber la fila exacta de cada agente.
    # Reconstruimos la misma lógica de lectura: recorrer rangos y buscar coincidencia de código.
    # Esto es más fiable que guardar la fila en la BD, pero podemos hacerlo dinámicamente.
    
    # Recorrer cada zona y sus filas
    for zona, (fila_inicio, fila_fin) in RANGOS_POR_ZONA.items():
        # Filtrar agentes de esta zona
        agentes_zona = df_agentes[df_agentes["zona"] == zona]
        if agentes_zona.empty:
            continue
        
        # Recorrer las filas del rango
        for fila_idx in range(fila_inicio - 1, min(fila_fin, ws.max_row)):
            # Leer código de la celda columna D (índice 4 en openpyxl es D)
            codigo_celda = ws.cell(row=fila_idx+1, column=4).value
            if codigo_celda is None:
                continue
            codigo_str = str(codigo_celda).strip()
            # Buscar agente con ese código en esta zona
            agente = agentes_zona[agentes_zona["codigo"] == codigo_str]
            if agente.empty:
                continue
            agente_id = agente.iloc[0]["id"]
            # Obtener turnos actuales desde la BD
            turnos = db.get_turnos(agente_id)
            # Escribir en las columnas de turno (columna 6,8,... para cada día)
            for dia in range(1, 32):
                col_turno = 5 + (dia-1)*2 + 1  # openpyxl 1-indexed: la primera columna de turno es la 6 (E? no, la 6 es F). Revisemos:
                # En pandas, columna índice 5 (0-index) corresponde a la columna 6 en Excel (F).
                # Día 1: col 5+0*2 =5 (0-index) -> col 6 (F). Día 2: col 7 -> col 8 (H), etc.
                # openpyxl: columna 6 es F, 8 es H...
                col_num = 6 + (dia-1)*2
                valor_turno = turnos[dia-1]
                if valor_turno:
                    ws.cell(row=fila_idx+1, column=col_num, value=valor_turno)
                else:
                    ws.cell(row=fila_idx+1, column=col_num, value="0")  # mantener 0 para vacío
    
    # Guardar en un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============================================================
# VISTA CALENDARIO (colores por turno)
# ============================================================

def mostrar_calendario(turnos):
    """Devuelve HTML del calendario para un agente, con colores según el turno."""
    dias_semana = ["LUN", "MAR", "MIÉ", "JUE", "VIE", "SÁB", "DOM"]
    primer_dia_semana = 4  # 1 de mayo 2026 es viernes (0=lunes, 4=viernes)
    
    # Mapa de colores para los turnos (ajústalo según tus turnos reales)
    color_turno = {
        "M": "#DCFCE7",   # verde claro
        "T": "#FFEDD5",   # naranja claro
        "N": "#F3F4F6",   # gris claro
        "DESCANSO": "#E0E7FF", # azul muy claro
        "LIBRE": "#E0E7FF"
    }
    color_default = "#FEF9C3"  # amarillo pálido
    
    html = '<table style="width:100%; border-collapse:collapse; text-align:center;">'
    # Cabecera
    html += '<thead><tr style="background-color:#1e293b; color:white;">'
    for dia in dias_semana:
        html += f'<th style="padding:8px; border:1px solid #475569; font-size:0.75rem;">{dia}</th>'
    html += '</tr></thead><tbody>'
    
    # Calendario (5 semanas)
    for semana in range(5):
        html += '<tr>'
        for dia_semana in range(7):
            dia_num = semana * 7 + dia_semana - primer_dia_semana + 1
            if 1 <= dia_num <= 31:
                turno = turnos[dia_num - 1] if turnos[dia_num - 1] else ""
                # Elegir color según turno
                bg_color = color_turno.get(turno, color_default)
                # Si no hay turno, usar colores suaves por paridad
                if turno == "":
                    bg_color = "#DBEAFE" if dia_num % 2 == 0 else "#FEF3C7"
                html += f'''
                <td style="background-color:{bg_color}; padding:8px 4px; border:1px solid #cbd5e1;">
                    <div style="font-weight:bold; font-size:0.7rem;">{dia_num}</div>
                    <div style="font-size:0.75rem; font-weight:500;">{turno if turno else "—"}</div>
                </td>
                '''
            else:
                html += '<td style="background-color:#f1f5f9; border:1px solid #cbd5e1;">&nbsp;</td>'
        html += '</tr>'
    html += '</tbody></table>'
    return html


# ============================================================
# INTERFAZ PRINCIPAL
# ============================================================

if 'db' not in st.session_state:
    st.session_state.db = GestorDB()
if 'datos_cargados' not in st.session_state:
    st.session_state.datos_cargados = False
if 'archivo_cargado' not in st.session_state:
    st.session_state.archivo_cargado = None
if 'mensaje' not in st.session_state:
    st.session_state.mensaje = None

with st.sidebar:
    st.header("📁 Cargar Excel")
    archivo = st.file_uploader("Selecciona AÑO 2026 ESTACIONES .xlsx", type=["xlsx"])
    
    if archivo:
        if st.button("📥 CARGAR AGENTES", type="primary", use_container_width=True):
            with st.spinner("Cargando agentes desde Excel..."):
                # Guardar temporalmente
                with open("temp_original.xlsx", "wb") as f:
                    f.write(archivo.getbuffer())
                # Cargar datos
                agentes_por_zona = cargar_por_rangos("temp_original.xlsx")
            
            total = sum(len(ag) for ag in agentes_por_zona.values())
            if total > 0:
                st.session_state.db.limpiar()
                for zona, agentes in agentes_por_zona.items():
                    for ag in agentes:
                        agente_id = st.session_state.db.guardar_agente(ag["codigo"], ag["nombre"], zona)
                        st.session_state.db.guardar_turnos(agente_id, ag["turnos"])
                st.session_state.datos_cargados = True
                st.session_state.archivo_cargado = "temp_original.xlsx"
                st.session_state.mensaje = f"✅ Cargados {total} agentes. Recuerda guardar cambios si modificas turnos."
                st.rerun()
            else:
                st.error("No se encontraron agentes válidos (código vacío o 0, o nombre con DESPLAZADO/VACANTE)")
    
    if st.session_state.datos_cargados:
        st.markdown("---")
        df_agentes = st.session_state.db.get_agentes()
        zonas = df_agentes["zona"].unique()
        zona_sel = st.selectbox("📍 Zona", ["TODAS"] + list(zonas))
        st.session_state.zona_sel = zona_sel
        st.metric("Total agentes", len(df_agentes))
        
        # Botón para guardar cambios en el Excel original
        if st.button("💾 Guardar cambios en Excel original", use_container_width=True):
            if st.session_state.archivo_cargado:
                try:
                    nuevo_excel = guardar_cambios_en_excel(st.session_state.archivo_cargado, st.session_state.db)
                    st.download_button(
                        label="📥 Descargar Excel actualizado",
                        data=nuevo_excel,
                        file_name="CUADRANTE_MODIFICADO_MAYO2026.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                    st.success("Excel actualizado listo para descargar. Guárdalo en tu equipo.")
                except Exception as e:
                    st.error(f"Error al generar el Excel: {e}")
            else:
                st.warning("No hay archivo original cargado.")

# Mostrar mensaje
if st.session_state.mensaje:
    if "✅" in st.session_state.mensaje:
        st.success(st.session_state.mensaje)
    else:
        st.info(st.session_state.mensaje)
    st.session_state.mensaje = None

if not st.session_state.datos_cargados:
    st.info("👈 **Carga el archivo Excel y haz clic en CARGAR AGENTES**")
    with st.expander("📖 Configuración"):
        st.markdown("""
        | Zona | Filas |
        |------|-------|
        | **JC** | 12 - 127 |
        | **AEZ6** | 140 - 181 |
        | **AEZ7** | 196 - 245 |
        | **AEZ8** | 262 - 309 |
        
        **No se cargan agentes con:**
        - Código vacío o igual a `0`
        - Nombre que contenga `DESPLAZADO` o `VACANTE`
        
        **Turnos:** se conserva el texto exacto del Excel (ej. "M", "T", "N", "DESCANSO"...).  
        Los turnos `0` o `0.0` se tratan como vacíos.
        """)
else:
    zona_sel = st.session_state.get('zona_sel', 'TODAS')
    df_agentes = st.session_state.db.get_agentes()
    if zona_sel != "TODAS":
        df_agentes = df_agentes[df_agentes["zona"] == zona_sel]
    
    st.markdown(f"## 📊 {zona_sel}")
    st.caption(f"{len(df_agentes)} agentes | 🎨 Colores: verde=M, naranja=T, gris=N, amarillo=otros, azul/ámbar=sin turno")
    
    # Grid de 2 columnas
    cols = st.columns(2)
    for idx, (_, agente) in enumerate(df_agentes.iterrows()):
        with cols[idx % 2]:
            turnos = st.session_state.db.get_turnos(agente["id"])
            with st.expander(f"📌 {agente['codigo']} - {agente['nombre']}"):
                st.markdown(mostrar_calendario(turnos), unsafe_allow_html=True)
    
    # Intercambio de turnos mejorado
    st.markdown("---")
    st.markdown("## 🔄 Intercambiar turnos")
    if len(df_agentes) >= 2:
        col1, col2, col3 = st.columns(3)
        with col1:
            nombres = [f"{row['codigo']} - {row['nombre']}" for _, row in df_agentes.iterrows()]
            ag1 = st.selectbox("Agente 1", nombres, key="ag1")
            idx1 = nombres.index(ag1)
            ag1_id = df_agentes.iloc[idx1]["id"]
            ag1_nombre = df_agentes.iloc[idx1]["nombre"]
            ag1_cod = df_agentes.iloc[idx1]["codigo"]
        with col2:
            ag2 = st.selectbox("Agente 2", nombres, key="ag2")
            idx2 = nombres.index(ag2)
            ag2_id = df_agentes.iloc[idx2]["id"]
            ag2_nombre = df_agentes.iloc[idx2]["nombre"]
            ag2_cod = df_agentes.iloc[idx2]["codigo"]
        with col3:
            dia = st.selectbox("Día", list(range(1, 32)), key="dia")
        
        turno1 = st.session_state.db.get_turno(ag1_id, dia)
        turno2 = st.session_state.db.get_turno(ag2_id, dia)
        
        st.info(f"📌 **Turnos actuales del día {dia}:**")
        st.write(f"   • {ag1_cod} - {ag1_nombre}: **`{turno1 if turno1 else '—'}`**")
        st.write(f"   • {ag2_cod} - {ag2_nombre}: **`{turno2 if turno2 else '—'}`**")
        
        if st.button("🔄 INTERCAMBIAR TURNOS", type="primary", use_container_width=True):
            t1, t2 = st.session_state.db.intercambiar_turnos(ag1_id, ag2_id, dia)
            st.balloons()
            st.session_state.mensaje = f"✅ Turnos intercambiados: {ag1_cod} ↔ {ag2_cod} (día {dia}). Ahora {ag1_cod} tiene '{t2 if t2 else '—'}' y {ag2_cod} tiene '{t1 if t1 else '—'}'."
            st.rerun()
    else:
        st.warning("Se necesitan al menos 2 agentes para intercambiar turnos.")
    
    # Botón para recargar todo (limpiar y volver a cargar otro archivo)
    if st.button("🔄 Cambiar archivo / Recargar", use_container_width=True):
        st.session_state.datos_cargados = False
        st.session_state.archivo_cargado = None
        st.rerun()
